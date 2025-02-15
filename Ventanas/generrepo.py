import os
import sys
import sqlite3
from collections import defaultdict
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


def resumenfinal():
    def obtener_datos():
        dirantbase = os.path.abspath(os.path.join(os.path.dirname(__file__),'..'))
        dirbasededatos = os.path.join(dirantbase,'DB')
        pathdb = os.path.join(dirbasededatos,'db')
        co = sqlite3.connect(pathdb)
        cur = co.cursor()
        consulta = '''
        SELECT 
            e.id AS estudiante_id,
            e.nombres,
            e.apellidos,
            e.ci_estudiante,
            e.cedula_escolar,
            e.Pasaporte,
            e.edad,
            a.anio_a_cursar,
            m.abreviatura AS materia_abrev,
            n.lapso,
            n.nota
        FROM estudiantes e
        JOIN academicos a ON e.id = a.estudiante_id
        JOIN estudiante_materias em ON e.id = em.estudiante_id
        JOIN materias m ON em.materia_id = m.id
        LEFT JOIN notas n ON e.id = n.estudiante_id AND m.id = n.materia_id
        ORDER BY a.anio_a_cursar, e.apellidos, e.nombres, m.abreviatura, n.lapso
        '''
        cur.execute(consulta)
        resultados = cur.fetchall()
        co.close()
        return resultados

    def procesar_datos(resultados):
        datos_por_anio = defaultdict(lambda: defaultdict(dict))

        for fila in resultados:
            (estudiante_id, nombres, apellidos, ci_estudiante, cedula_escolar,
            pasaporte, edad, anio, materia_abrev, lapso, nota) = fila

            # Crear una clave única para cada estudiante
            clave_estudiante = (estudiante_id, nombres, apellidos, ci_estudiante,
                                cedula_escolar, pasaporte, edad)

            # Agregar los datos al diccionario
            estudiante = datos_por_anio[anio].setdefault(clave_estudiante, {
                'Nombre': nombres,
                'Apellido': apellidos,
                'CI': ci_estudiante,
                'CI Escolar': cedula_escolar,
                'Pasaporte': pasaporte,
                'Edad': edad,
                'Materias': defaultdict(dict)
            })

            if materia_abrev:
                if lapso is not None and nota is not None:
                    estudiante['Materias'][materia_abrev][lapso] = nota
                else:
                    # Aseguramos que la materia esté presente
                    if materia_abrev not in estudiante['Materias']:
                        estudiante['Materias'][materia_abrev] = {}

        return datos_por_anio

    def crear_directorio_reporte():
        dir_actual = os.path.dirname(os.path.abspath(__file__))
        dir_test1 = os.path.abspath(os.path.join(dir_actual, '..'))
        dir_reportes = os.path.join(dir_test1, 'reportes')
        if not os.path.exists(dir_reportes):
            os.makedirs(dir_reportes)
        ahora = datetime.now()
        fecha_hora_formateada = ahora.strftime('%d-%m-%Y-%H-%M')
        dir_fecha_hora = os.path.join(dir_reportes, fecha_hora_formateada)
        os.makedirs(dir_fecha_hora)
        return dir_fecha_hora

    def generar_reporte(datos_por_anio):
        # Crear el directorio donde se guardará el reporte
        ruta_reporte = crear_directorio_reporte()
        ruta_archivo = os.path.join(ruta_reporte, 'Resumen Final.xlsx')

        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as escritor:
            hojas_creadas = False
            for anio, estudiantes in datos_por_anio.items():
                if not estudiantes:
                    continue

                filas = []

                # Preparar una lista de todas las materias para el año actual
                materias_set = set()
                for datos_estudiante in estudiantes.values():
                    materias_set.update(datos_estudiante['Materias'].keys())
                materias_ordenadas = sorted(materias_set)

                for datos_clave, datos_estudiante in estudiantes.items():
                    fila = {
                        'Nombres': datos_estudiante['Nombre'],
                        'Apellidos': datos_estudiante['Apellido'],
                        'CI': datos_estudiante['CI'],
                        'CI Escolar': datos_estudiante['CI Escolar'],
                        'Pasaporte': datos_estudiante['Pasaporte'],
                        'Edad': datos_estudiante['Edad']
                    }

                    for materia in materias_ordenadas:
                        notas = datos_estudiante['Materias'].get(materia, {})
                        # Agregar las notas de cada lapso
                        fila[f'{materia} - 1er Lapso'] = notas.get(1, '')
                        fila[f'{materia} - 2do Lapso'] = notas.get(2, '')
                        fila[f'{materia} - 3er Lapso'] = notas.get(3, '')

                        # Calcular la nota definitiva
                        notas_validas = [nota for lapso, nota in notas.items() if nota is not None]
                        if notas_validas:
                            definitiva = sum(notas_validas) / 3
                            fila[f'{materia} - Definitiva'] = round(definitiva, 2)
                        else:
                            fila[f'{materia} - Definitiva'] = ''

                    filas.append(fila)

                # Convertir la lista de filas en un DataFrame
                df = pd.DataFrame(filas)

                if df.empty:
                    continue

                # Ordenar las columnas
                columnas_iniciales = ['Nombres', 'Apellidos', 'CI', 'CI Escolar', 'Pasaporte', 'Edad']
                columnas_materias = []

                for materia in materias_ordenadas:
                    columnas_materias.extend([
                        f'{materia} - 1er Lapso',
                        f'{materia} - 2do Lapso',
                        f'{materia} - 3er Lapso',
                        f'{materia} - Definitiva'
                    ])

                columnas_ordenadas = columnas_iniciales + columnas_materias
                df = df[columnas_ordenadas]

                # DataFrame en una hoja del Excel
                nombre_hoja = f'{anio} año'  # Por ejemplo, '1er año', '2do año'
                df.to_excel(escritor, sheet_name=nombre_hoja, index=False)

                # Obtener el workbook y la hoja
                workbook = escritor.book
                worksheet = workbook[nombre_hoja]

                ajustar_formato_excel(worksheet, df)

                hojas_creadas = True

            if not hojas_creadas:
                print("No se encontraron datos para generar el reporte.")
                return

            print(f"Reporte generado exitosamente en: {ruta_archivo}")

    def ajustar_formato_excel(worksheet, df):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font

        # Fusionar celdas para el título del año
        max_col = len(df.columns)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        celda_titulo = worksheet.cell(row=1, column=1)
        celda_titulo.value = worksheet.title
        celda_titulo.alignment = Alignment(horizontal='center')
        celda_titulo.font = Font(bold=True, size=14)

        # Repetir encabezados después del título
        worksheet.insert_rows(2)
        for col_num, column_title in enumerate(df.columns, 1):
            celda_encabezado = worksheet.cell(row=2, column=col_num)
            celda_encabezado.value = column_title
            celda_encabezado.font = Font(bold=True)
            celda_encabezado.alignment = Alignment(horizontal='center', wrap_text=True)

        # Ajustar ancho de columnas
        for i, column in enumerate(df.columns, 1):
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
            worksheet.column_dimensions[get_column_letter(i)].width = max_length

        # Ajustar alineación de datos
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    resultados1 = obtener_datos()
    if not resultados1:
        print("No se encontraron datos en la consulta.")
    else:
        datos_por_anio = procesar_datos(resultados1)
        generar_reporte(datos_por_anio)

def boletin():
    def obtener_datos_boletin():
        dirantbase = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        dirbasededatos = os.path.join(dirantbase, 'DB')
        pathdb = os.path.join(dirbasededatos, 'db')
        co = sqlite3.connect(pathdb)
        cur = co.cursor()
        consulta = '''
        SELECT 
            e.id AS estudiante_id,
            e.nombres,
            e.apellidos,
            e.ci_estudiante,
            e.cedula_escolar,
            e.Pasaporte,
            e.edad,
            a.anio_a_cursar,
            m.nombre AS materia_nombre,
            n.lapso,
            n.nota
        FROM estudiantes e
        JOIN academicos a ON e.id = a.estudiante_id
        JOIN estudiante_materias em ON e.id = em.estudiante_id
        JOIN materias m ON em.materia_id = m.id
        LEFT JOIN notas n ON e.id = n.estudiante_id AND m.id = n.materia_id AND n.lapso IS NOT NULL
        ORDER BY a.anio_a_cursar, e.apellidos, e.nombres, m.nombre, n.lapso
        '''
        cur.execute(consulta)
        resultados = cur.fetchall()
        co.close()
        return resultados

    def procesar_datos_boletin(resultados):
        datos_estudiantes = {}

        for fila in resultados:
            (estudiante_id, nombres, apellidos, ci_estudiante, cedula_escolar,
            pasaporte, edad, anio, materia_nombre, lapso, nota) = fila

            clave_estudiante = estudiante_id

            if clave_estudiante not in datos_estudiantes:
                # Determinar qué documento utilizar
                documento = ci_estudiante or cedula_escolar or pasaporte or ''
                datos_estudiantes[clave_estudiante] = {
                    'Nombre Completo': f"{nombres} {apellidos}",
                    'Documento': documento,
                    'Edad': edad,
                    'Año': anio,
                    'Materias': {}
                }

            if materia_nombre not in datos_estudiantes[clave_estudiante]['Materias']:
                datos_estudiantes[clave_estudiante]['Materias'][materia_nombre] = {
                    '1er Lapso': '',
                    '2do Lapso': '',
                    '3er Lapso': '',
                    'Cal. Def.': ''
                }

            # Agregar la nota al lapso correspondiente
            if lapso == 1:
                datos_estudiantes[clave_estudiante]['Materias'][materia_nombre]['1er Lapso'] = nota or ''
            elif lapso == 2:
                datos_estudiantes[clave_estudiante]['Materias'][materia_nombre]['2do Lapso'] = nota or ''
            elif lapso == 3:
                datos_estudiantes[clave_estudiante]['Materias'][materia_nombre]['3er Lapso'] = nota or ''

        # Calcular la nota definitiva para cada materia
        for estudiante in datos_estudiantes.values():
            for materia, notas in estudiante['Materias'].items():
                notas_lapsos = [notas['1er Lapso'], notas['2do Lapso'], notas['3er Lapso']]
                notas_validas = [n for n in notas_lapsos if isinstance(n, (int, float))]
                if notas_validas:
                    cal_def = sum(notas_validas) / len(notas_validas)
                    notas['Cal. Def.'] = round(cal_def, 2)
                else:
                    notas['Cal. Def.'] = ''

        return datos_estudiantes

    def crear_directorio_reporte():
        dir_actual = os.path.dirname(os.path.abspath(__file__))
        dir_test1 = os.path.abspath(os.path.join(dir_actual, '..'))
        dir_reportes = os.path.join(dir_test1, 'reportes')
        if not os.path.exists(dir_reportes):
            os.makedirs(dir_reportes)
        ahora = datetime.now()
        fecha_hora_formateada = ahora.strftime('%d-%m-%Y-%H-%M')
        dir_fecha_hora = os.path.join(dir_reportes, fecha_hora_formateada)
        os.makedirs(dir_fecha_hora)
        return dir_fecha_hora

    def generar_boletines(datos_estudiantes):

        # Crear el directorio donde se guardará el reporte
        ruta_reporte = crear_directorio_reporte()
        ruta_archivo = os.path.join(ruta_reporte, 'Boletines.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "Boletines"

        # Configurar el diseño de página
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        ws.page_margins.left = ws.page_margins.right = 0.5

        # Definir estilos
        titulo_estilo = Font(bold=True, size=14)
        encabezado_estilo = Font(bold=True)
        alineación_centrada = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

        # Variables para controlar la posición en la hoja
        fila_actual = 1

        for estudiante_id, datos in datos_estudiantes.items():
            # Agregar información del estudiante
            ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=9)
            celda_titulo = ws.cell(row=fila_actual, column=1)
            celda_titulo.value = f"Boletín de Calificaciones - {datos['Año']}° Año"
            celda_titulo.font = titulo_estilo
            celda_titulo.alignment = alineación_centrada
            fila_actual += 2

            ws.cell(row=fila_actual, column=1, value="Estudiante:")
            ws.cell(row=fila_actual, column=2, value=datos['Nombre Completo'])
            ws.cell(row=fila_actual, column=4, value="Cédula de Identidad:")
            ws.cell(row=fila_actual, column=5, value=datos['Documento'])
            fila_actual += 2

            # Encabezados de la tabla de notas
            encabezados = [
                "Área de Formación", "1er Lapso", "2do Lapso", "3er Lapso",
                "Cal. Def.", "Observaciones"
            ]
            ws.append(encabezados)
            for col in range(1, len(encabezados) + 1):
                celda = ws.cell(row=fila_actual, column=col)
                celda.font = encabezado_estilo
                celda.alignment = alineación_centrada
                celda.border = borde
            fila_actual += 1

            # Agregar notas de las materias
            materias = datos['Materias']
            for materia, notas in materias.items():
                fila_materia = [
                    materia,
                    notas.get('1er Lapso', ''),
                    notas.get('2do Lapso', ''),
                    notas.get('3er Lapso', ''),
                    notas.get('Cal. Def.', ''),
                    '',  # Observaciones
                ]
                ws.append(fila_materia)
                for col in range(1, len(fila_materia) + 1):
                    celda = ws.cell(row=fila_actual, column=col)
                    celda.alignment = alineación_centrada
                    celda.border = borde
                fila_actual += 1

            # Agregar el promedio general del estudiante
            promedios = [
                notas['Cal. Def.'] for notas in materias.values()
                if isinstance(notas['Cal. Def.'], (int, float))
            ]
            if promedios:
                promedio_general = round(sum(promedios) / len(promedios), 2)
            else:
                promedio_general = ''

            ws.cell(row=fila_actual, column=1, value="Promedio General:")
            ws.cell(row=fila_actual, column=2, value=promedio_general)
            fila_actual += 2

            # Insertar Salto de Página
            ws.row_breaks.append(Break(id=fila_actual - 1))

        # Ajustar anchos de columnas
        for col in ws.columns:
            max_length = 0
            col_idx = col[0].column  
            column = get_column_letter(col_idx)  
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(ruta_archivo)
        print(f"Boletines generados exitosamente en: {ruta_archivo}")

    resultados = obtener_datos_boletin()
    if not resultados:
        print("No se encontraron datos para generar los boletines.")
    else:
        datos_estudiantes = procesar_datos_boletin(resultados)
        generar_boletines(datos_estudiantes)

def bdatos():
    
    def obtener_datos_estudiantes():
        dirantbase = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        dirbasededatos = os.path.join(dirantbase, 'DB')
        pathdb = os.path.join(dirbasededatos, 'db')
        co = sqlite3.connect(pathdb)
        cur = co.cursor()

        consulta = '''
        SELECT 
            e.id AS estudiante_id,
            e.nombres,
            e.apellidos,
            e.fecha_nacimiento,
            e.edad,
            e.sexo,
            e.lugar_nacimiento,
            e.nacionalidad,
            e.Pasaporte,
            e.ci_estudiante,
            e.cedula_escolar,
            e.estatura,
            e.peso,
            e.talla_zapato,
            e.talla_camisa,
            e.talla_pantalon,
            e.enfermedad_cronica,
            e.observaciones,
            c.direccion,
            c.telefono_movil,
            c.email,
            a.anio_a_cursar,
            a.repite
        FROM estudiantes e
        LEFT JOIN estudiantes_contactos ec ON e.id = ec.estudiante_id
        LEFT JOIN contactos c ON ec.contacto_id = c.id
        LEFT JOIN academicos a ON e.id = a.estudiante_id
        ORDER BY e.apellidos, e.nombres
        '''
        cur.execute(consulta)
        resultados = cur.fetchall()
        co.close()
        return resultados

    def generar_reporte_estudiantes(datos):
        ruta_reporte = crear_directorio_reporte()
        ruta_archivo = os.path.join(ruta_reporte, 'Reporte Estudiantes.xlsx')

        columnas = ['ID Estudiante', 'Nombres', 'Apellidos', 'Fecha de Nacimiento', 'Edad', 'Sexo', 
                    'Lugar de Nacimiento', 'Nacionalidad', 'Pasaporte', 'Cédula Estudiante', 'Cédula Escolar',
                    'Estatura', 'Peso', 'Talla de Zapato', 'Talla de Camisa', 'Talla de Pantalón',
                    'Enfermedad Crónica', 'Observaciones', 'Dirección', 'Teléfono Móvil', 'Email',
                    'Año a Cursar', 'Repite']

        df = pd.DataFrame(datos, columns=columnas)
        # Eliminar duplicados si es necesario
        df = df.drop_duplicates(subset=['ID Estudiante'])
        # Reordenamos las columnas si es necesario
        df = df[columnas]

        # Guardado del dataframe
        df.to_excel(ruta_archivo, index=False, sheet_name='Estudiantes')

        print(f"Reporte generado exitosamente en: {ruta_archivo}")

    def crear_directorio_reporte():
        dir_actual = os.path.dirname(os.path.abspath(__file__))
        dir_test1 = os.path.abspath(os.path.join(dir_actual, '..'))
        dir_reportes = os.path.join(dir_test1, 'reportes')
        if not os.path.exists(dir_reportes):
            os.makedirs(dir_reportes)
        ahora = datetime.now()
        fecha_hora_formateada = ahora.strftime('%d-%m-%Y-%H-%M')
        dir_fecha_hora = os.path.join(dir_reportes, fecha_hora_formateada)
        os.makedirs(dir_fecha_hora)
        return dir_fecha_hora

    datos_estudiantes = obtener_datos_estudiantes()
    if not datos_estudiantes:
        print("No se encontraron datos en la base de datos.")
    else:
        generar_reporte_estudiantes(datos_estudiantes)

'''
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººººSistema de gestion Academica Realizadoºººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººººPor los estudiantes de 4to Semestre enºººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººº     Analisis y diseño de Sistemas    ºººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººº Alejandra Brito 30414509 04248613003 ºººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººº Edinson Lozano 31389068 04248487046  ºººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººº Victor Gonzalez 29985491 04161942133 ºººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººº Leonardo Calojero 30729451 04121959745ºººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººº Enmanuel Carreño 31177838 04128566093 ºººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººº Kliver Rivas 30685024     04162917106 ºººººººººººººººººººººººººººººººººººººººº
ººººººººººººººººººººººººººººººººº Github: https://github.com/fireclik/sc ººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº
ºººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººººº

'''
    